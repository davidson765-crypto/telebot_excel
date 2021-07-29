import telebot
from telebot import types
import xlsxwriter
from openpyxl import load_workbook
import pandas as pd


bot = telebot.TeleBot('1823265709:AAEvLJzuy8_fQpAAlOzlivRhUbOOfFDlKKk')

@bot.message_handler(content_types=['text'])


def get_text_messages(message):


	for i in message.text:
		if i not in words_list:
			pas = True
		else:
			pas = False

	if len(message.text)!=11 and pas==False:
		bot.send_message(message.from_user.id, 'Здраствуйте, напишите мне свой номер.')
	elif len(message.text)==11 and pas==True:
	
		wb = load_workbook('team.xlsx')
		ws = wb.active
		sheet_ranges = wb['Лист1']

		for i in range(1,11):
			if str(sheet_ranges[f'A{i}'].value)==message.text:
				padan = i
				pas = True
				break
			else:
				pas = False
		if pas == True:
			print(padan)
			bot.send_message(message.from_user.id, 'Напишите информацию о номере (например ФИО).')
			bot.register_next_step_handler(message,rigister,padan)
		else:
			bot.send_message(message.from_user.id, 'Такого номера нет в базе.')
				

def rigister(message,padan):

		wb = load_workbook('team.xlsx')
		ws = wb.active
		sheet_ranges = wb['Лист1']

		ws[f'B{padan}']=message.text

		wb.save('team.xlsx')

		bot.send_message(message.from_user.id, 'Thапись сделана, хорошего дня!')



words_list = list('qwertyuiopasdfghjklzxcvbnm')

bot.polling(none_stop=True, interval=0)


